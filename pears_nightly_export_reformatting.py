import boto3
import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

# An AWS named profile will need to be created for accessing the PEARS S3
# without additional modifications to this script.
# 
# Obtain the key and secret from PEARS support.
# 
# After installing AWS CLI, run the following command:
# aws configure --profile your_profile_name
# 
# Reference:
# https://docs.aws.amazon.com/cli/latest/userguide/cli-configure-profiles.html


# Use PEARS AWS S3 credentials
session = boto3.Session(profile_name='your_profile_name')  

# Illinois Extension is only authorized to access the following subdirectory
prefix = 'uie/'
# Access S3 objects uploaded the day reformatting script is run
ts = pd.to_datetime("today").strftime("%Y/%m/%d/")
conn = session.client('s3')
my_bucket = 'exports.pears.oeie.org'

response = conn.list_objects_v2(
    Bucket=my_bucket,
    Prefix=prefix + ts,
    MaxKeys=100)

# Define the target directory for PEARS imports
os.chdir(
    r"C:\Users\netid\your\target\directory")

# Download the Excel files to the target directory

for f in response['Contents']:
    file = f['Key']
    filename = file[file.rfind('/') + 1:]
    conn.download_file(my_bucket, file, filename)

# Desired PEARS modules to reformat
# 'Excel_File', 'Sheet Name'

import_modules = [['Program_Activities', 'Program Activity Data'],
                  ['Indirect_Activity', 'Indirect Activity Data'],
                  ['Coalition', 'Coalition Data'],
                  ['Partnership', 'Partnership Data'],
                  ['PSE_Site_Activity', 'PSE Data']]

# Custom fields that require reformatting
# Reformatting only needed for multi-select dropdowns

custom_field_labels = ['fcs_program_team', 'snap_ed_grant_goals', 'fcs_grant_goals', 'fcs_special_projects',
                       'snap_ed_special_projects']

# Function to convert custom field's label to its dropdown value
# text: string value of the label suffix
# custom_field_label: string for the custom field label's prefix
def replace_all(text, custom_field_label):
    dic = {
        custom_field_label: '',
        # Map label suffixes to response options
        'family_life': 'Family Life',
        'nutrition_wellness': 'Nutrition & Wellness',
        'consumer_economics': 'Consumer Economics',
        'snap_ed': 'SNAP-Ed',
        'efnep': 'EFNEP',
        'improve_diet_quality': 'Improve diet quality',
        'increase_physical_activity_opportunities': 'Increase physical activity opportunities',
        'increase_food_access': 'Increase food access',
        'none': 'None',
        'abcs_of_school_nutrition': 'ABCs of School Nutrition',
        'growing_together_illinois': 'Growing Together Illinois',
        'heat': 'HEAT',
        'cphp_shape_up_chicago_youth_trainers': 'CPHP - Shape Up Chicago Youth Trainers',
        'cphp_chicago_grows_groceries': 'CPHP - Chicago Grows Groceries',
        '_': '',
    }
    for i, j in dic.items():
        text = text.replace(i, j)
    return text

# In-place function to convert custom field value binary columns into a single custom field column of list-like strings
# records: dataframe to reformat
# labels: list of custom labels to iterate through
def reformat(records, labels):
    for label in labels:
        binary_cols = records.columns[records.columns.str.contains(label)]
        if binary_cols.empty:
            continue
        for col in binary_cols:
            records.loc[records[col] == 1, col] = replace_all(col, label)
            records.loc[records[col] == 0, col] = ''
        # Create custom field column of list-like strings
        records[label] = records[binary_cols].apply(lambda row: ','.join(row.values.astype(str)), axis=1).str.strip(
            ',').str.replace(r',+', ',', regex=True)
        records.loc[records[label] == '', label] = np.nan
        # Remove custom field value binary columns
        records.drop(columns=binary_cols, inplace=True)

# Function to update workbook with reformatted data
# file_name: Module export excel file
# sheet: Excel sheet of module data
# dataframe: reformatted data to replace sheet
def write_excel(file_name, sheet, dataframe):
    # Open module's workbook
    book = openpyxl.load_workbook(file_name)
    # Reformat module data tab without importing every tab
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='w') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        ws = writer.sheets[sheet]
        # Clear original contents of sheet
        ws.delete_cols(1, ws.max_column)
        ws.delete_rows(1, ws.max_row)
        # Add reformatted data to sheet
        dataframe.to_excel(writer, sheet_name=sheet, index=False)
        # Use original sheet's style formatting
        font = Font(bold=False)
        fill = PatternFill(fill_type='solid', fgColor='00C0C0C0')
        border = Border(bottom=Side(border_style=None, color='FF000000'))
        for c in ws["1:1"]:
            c.font = font
            c.fill = fill
            c.border = border


# Full execution

for item in import_modules:
    src = os.getcwd() + "\\" + item[0] + "_Export.xlsx"
    # Read module's sheet
    df = pd.read_excel(src, sheet_name=item[1])
    # Remove custom data tag from column labels
    df.columns = df.columns.str.replace(r'_custom_data', '')
    reformat(df, custom_field_labels)
    # Overite each workbook with reformatted data
    write_excel(src, item[1], df)
    
